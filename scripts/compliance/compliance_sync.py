#!/usr/bin/env python3
"""
SwiftLoan — Compliance drift-detection & doc-sync engine.

WHAT IT DOES
------------
Runs a battery of deterministic control checks (the C1–C14 findings + key ISMS
controls) against the current codebase, then:
  1. Regenerates docs/compliance/COMPLIANCE-STATUS.md  (machine-maintained live
     control-status table, mapped to ISO 27001:2022 Annex A + SOC 2 TSC).
  2. Writes docs/compliance/compliance-status.csv       (always, dependency-free).
  3. Updates the "Live Control Status" sheet of the evidence-pack .xlsx
     (only if openpyxl is installed and the file exists).
  4. Detects DRIFT: if any monitored source file changed since the last sync
     (tracked in .code-manifest.json), it flags the affected controls and the
     compliance documents for human review.
  5. Auto-appends newly-merged pull requests (via `gh pr list`, or a `git log`
     merge-commit fallback) to the "Change Log" sheet of the SDLC tracker
     (07-*.xlsx) — idempotent: PRs already referenced in the sheet are skipped.

SCOPE / HONESTY
---------------
Machine-checkable control status (Implemented / Open) and the file:line evidence
are updated automatically. The narrative prose in the six documents is NOT
rewritten by a script — instead, drift is flagged with a clear "REVIEW REQUIRED"
list so a human (or an AI reviewer invoked by CI) updates the affected sections.
This avoids silently fabricating audit prose.

USAGE
-----
  python3 scripts/compliance/compliance_sync.py --update   # run + write artifacts
  python3 scripts/compliance/compliance_sync.py --check    # CI gate: exit 1 on undocumented drift
  python3 scripts/compliance/compliance_sync.py            # dry-run (print only)

Runs on push to main/master via .github/workflows/compliance-sync.yml.
"""
from __future__ import annotations
import json, os, re, subprocess, sys
from datetime import datetime, timezone
from pathlib import Path

REPO = Path(__file__).resolve().parents[2]
DOCS = REPO / "docs" / "compliance"
MANIFEST = DOCS / ".code-manifest.json"
STATUS_MD = DOCS / "COMPLIANCE-STATUS.md"
STATUS_CSV = DOCS / "compliance-status.csv"
EVIDENCE_XLSX = DOCS / "06-Compliance-Evidence-Pack-and-Claims-Matrix.xlsx"
SDLC_XLSX = DOCS / "07-SDLC-Change-Management-Tracker.xlsx"

THE_SIX_DOCS = [
    "01-Product-Requirements-Document",
    "02-Technical-Architecture-Document",
    "03-Security-and-Compliance-Document",
    "04-Test-Cases-Document",
    "05-Statement-of-Applicability",
    "06-Compliance-Evidence-Pack-and-Claims-Matrix",
    "07-SDLC-Change-Management-Tracker",
    "08-ISMS-Manual-Policies-and-Governance",
    "09-Security-and-Privacy-by-Design-Principles",
    "10-Certification-Readiness-and-Gap-Assessment",
]

def read(rel: str) -> str:
    p = REPO / rel
    try:
        return p.read_text(encoding="utf-8", errors="replace")
    except FileNotFoundError:
        return ""

def grep(rel: str, pattern: str, flags=re.I) -> bool:
    return re.search(pattern, read(rel), flags) is not None

# ── Control checks ────────────────────────────────────────────────────────────
# Each check returns True when the control is SATISFIED (risk closed).
# status: "Implemented" if satisfied else "Open".

def c1_pan_encrypted():
    s = read("server/prisma/schema.prisma")
    has_plain_pan = re.search(r"panNumber\s+String", s) is not None
    tokenized = re.search(r"panToken|panCipher|panEncrypted|@encrypted", s, re.I) is not None
    # closed when there's no plaintext panNumber column, or a tokenized/encrypted form exists
    return (not has_plain_pan) or tokenized

def c2_voice_redacts_sensitive():
    # closed when the page_context/read_screen path redacts value for sensitive fields
    s = read("src/voice/actionRegistry.ts") + read("src/voice/tools.ts")
    redacts = re.search(r"sensitive.*(redact|omit|\bnull\b|\[redacted\]|skip)", s, re.I) is not None
    return redacts

def c3_no_demo_otp_prod():
    # closed when render.yaml does NOT enable DEMO_LOGIN=true
    y = read("render.yaml")
    m = re.search(r"DEMO_LOGIN[\s\S]{0,60}?value:\s*['\"]?true", y, re.I)
    return m is None

def c4_no_default_admin():
    return not grep("server/prisma/seed.ws4.ts", r"admin123")

def c5_no_committed_key():
    # scan tracked files for an Ello-style key literal
    try:
        out = subprocess.run(["git", "grep", "-lE", r"ak_[A-Za-z0-9_-]{20,}\."],
                             cwd=REPO, capture_output=True, text=True)
        hits = [l for l in out.stdout.splitlines()
                if l and "voiceCredentials.local.example" not in l
                and "reference_md_files" not in l and "docs/compliance" not in l]
        return len(hits) == 0
    except Exception:
        return None  # unknown

def c6_secrets_fail_closed():
    return not grep("server/src/config/env.ts", r"'dev-access'|'dev-refresh'|\"dev-access\"|\"dev-refresh\"")

def c7_cors_locked():
    s = read("server/src/app.ts")
    open_cors = re.search(r"cors\(\s*\)", s) is not None
    return not open_cors

def c11_admin_token_httponly():
    return not grep("admin/src/lib/api.ts", r"localStorage")

def c13_voice_tls_only():
    return not grep("src/voice/config.ts", r"ws://|http://")

# Governance / CI security-automation artifacts (added this pass) — present => Implemented.
def _exists(rel): return (REPO / rel).exists()
def g_secret_scan():   return _exists(".github/workflows/security-scan.yml") and grep(".github/workflows/security-scan.yml", r"gitleaks")
def g_sast():          return grep(".github/workflows/security-scan.yml", r"codeql")
def g_sca():           return _exists(".github/dependabot.yml")
def g_codeowners():    return _exists("CODEOWNERS") or _exists(".github/CODEOWNERS")
def g_security_md():   return _exists("SECURITY.md")
def g_pr_template():   return _exists(".github/pull_request_template.md")
def g_license():       return _exists("LICENSE")

CHECKS = [
    dict(id="C1",  title="PAN encrypted/tokenized at rest (not plaintext)",
         iso="A.8.24, A.8.11", soc2="CC6.1, C1.1", fn=c1_pan_encrypted,
         evidence="server/prisma/schema.prisma:62,171",
         remediation="Tokenize/field-encrypt PAN; de-duplicate to one source."),
    dict(id="C2",  title="Voice page_context redacts sensitive field values",
         iso="A.8.11, A.5.34", soc2="C1.1, P4", fn=c2_voice_redacts_sensitive,
         evidence="src/voice/actionRegistry.ts:204; src/voice/screenGraph.ts:207",
         remediation="Redact sensitive values from page_context/read_screen."),
    dict(id="C3",  title="Fixed demo-OTP disabled in production",
         iso="A.8.5", soc2="CC6.1", fn=c3_no_demo_otp_prod,
         evidence="render.yaml DEMO_LOGIN; lib/crypto.ts:9-11",
         remediation="Set DEMO_LOGIN=false in prod; remove client bypass."),
    dict(id="C4",  title="No default/shared admin credential",
         iso="A.8.5, A.5.17", soc2="CC6.1", fn=c4_no_default_admin,
         evidence="server/prisma/seed.ws4.ts:74,82-83",
         remediation="Remove admin123 seed; force reset + MFA."),
    dict(id="C5",  title="No live API key committed to source",
         iso="A.8.24, A.5.19", soc2="CC6.1, CC9.2", fn=c5_no_committed_key,
         evidence="render.yaml:53; website/js/voice-widget.js:17",
         remediation="Rotate key; move to secrets manager; purge history."),
    dict(id="C6",  title="Secrets fail-closed (no dev fallback)",
         iso="A.8.9, A.8.24", soc2="CC6.1, CC8.1", fn=c6_secrets_fail_closed,
         evidence="server/src/config/env.ts:12-13",
         remediation="Remove dev-access/dev-refresh fallbacks; fail on boot."),
    dict(id="C7",  title="CORS restricted to allow-list",
         iso="A.8.20, A.8.23", soc2="CC6.6", fn=c7_cors_locked,
         evidence="server/src/app.ts:27",
         remediation="Replace open cors() with an origin allow-list."),
    dict(id="C11", title="Admin tokens not in localStorage (httpOnly)",
         iso="A.8.3, A.8.5", soc2="CC6.1", fn=c11_admin_token_httponly,
         evidence="admin/src/lib/api.ts:12-22",
         remediation="Move admin tokens to httpOnly cookies."),
    dict(id="C13", title="Voice transport TLS-only (no ws://,http://)",
         iso="A.8.24, A.8.20", soc2="CC6.7, C1.1", fn=c13_voice_tls_only,
         evidence="src/voice/config.ts",
         remediation="Force wss:// and https:// for voice transport."),
    dict(id="G5",  title="Secret scanning in CI (gitleaks)",
         iso="A.8.8, A.8.4", soc2="CC7.1", fn=g_secret_scan,
         evidence=".github/workflows/security-scan.yml",
         remediation="Keep gitleaks in CI; enable GitHub push protection."),
    dict(id="G6",  title="SAST in CI (CodeQL)",
         iso="A.8.25, A.8.28", soc2="CC8.1", fn=g_sast,
         evidence=".github/workflows/security-scan.yml",
         remediation="Keep CodeQL; triage findings."),
    dict(id="G7",  title="Dependency/SCA scanning (Dependabot + npm audit)",
         iso="A.8.8", soc2="CC7.1", fn=g_sca,
         evidence=".github/dependabot.yml",
         remediation="Keep Dependabot; act on advisories."),
    dict(id="G8",  title="Mandatory code-owner review",
         iso="A.8.4, A.8.32", soc2="CC8.1", fn=g_codeowners,
         evidence="CODEOWNERS",
         remediation="Enable 'Require Code Owner review' in branch protection."),
    dict(id="G9",  title="Vulnerability disclosure policy",
         iso="A.5.5, A.5.24", soc2="CC2.3", fn=g_security_md,
         evidence="SECURITY.md",
         remediation="Keep current; staff the security mailbox."),
    dict(id="G10", title="PR governance (security/compliance checklist)",
         iso="A.8.32", soc2="CC8.1", fn=g_pr_template,
         evidence=".github/pull_request_template.md",
         remediation="Keep; enforce checklist in review."),
    dict(id="G11", title="IP / licensing clarity",
         iso="A.5.32", soc2="CC1.1", fn=g_license,
         evidence="LICENSE",
         remediation="Keep proprietary notice current."),
]

# Files whose change should trigger a doc-review flag, mapped to control ids.
MONITORED = {
    "server/prisma/schema.prisma": ["C1", "data-model"],
    "server/src/config/env.ts": ["C6"],
    "server/src/app.ts": ["C7"],
    "render.yaml": ["C3", "C5"],
    "server/prisma/seed.ws4.ts": ["C4"],
    "website/js/voice-widget.js": ["C5"],
    "src/voice/actionRegistry.ts": ["C2"],
    "src/voice/screenGraph.ts": ["C2"],
    "src/voice/config.ts": ["C13"],
    "admin/src/lib/api.ts": ["C11"],
    "src/api/client.ts": ["C3", "transport"],
    "ios/SwiftLoan/Info.plist": ["transport"],
}

def sha(rel: str) -> str:
    import hashlib
    p = REPO / rel
    if not p.exists():
        return ""
    return hashlib.sha256(p.read_bytes()).hexdigest()[:16]

def git_head() -> str:
    try:
        return subprocess.run(["git", "rev-parse", "--short", "HEAD"], cwd=REPO,
                              capture_output=True, text=True).stdout.strip() or "unknown"
    except Exception:
        return "unknown"

def run_checks():
    rows = []
    for c in CHECKS:
        try:
            res = c["fn"]()
        except Exception as e:
            res = None
        status = "Implemented" if res is True else ("Open" if res is False else "Unknown")
        rows.append({**{k: c[k] for k in ("id", "title", "iso", "soc2", "evidence", "remediation")},
                     "status": status})
    return rows

def detect_drift():
    prev = {}
    if MANIFEST.exists():
        try:
            prev = json.loads(MANIFEST.read_text()).get("files", {})
        except Exception:
            prev = {}
    cur, changed = {}, []
    for rel in MONITORED:
        cur[rel] = sha(rel)
        if prev.get(rel, "__none__") != cur[rel] and prev:  # skip first-run noise
            changed.append(rel)
    affected = sorted({cid for rel in changed for cid in MONITORED[rel]})
    return cur, changed, affected

def write_status_md(rows, changed, affected, head, stamp):
    openc = [r for r in rows if r["status"] == "Open"]
    lines = []
    lines.append("# SwiftLoan — Live Compliance Control Status (auto-generated)\n")
    lines.append(f"> Generated by `scripts/compliance/compliance_sync.py` — **do not edit by hand.**  ")
    lines.append(f"> Commit `{head}` · {stamp} · {len(rows)-len(openc)}/{len(rows)} controls Implemented\n")
    lines.append("| ID | Control | Status | ISO 27001:2022 | SOC 2 TSC | Evidence | Remediation |")
    lines.append("|----|---------|--------|----------------|-----------|----------|-------------|")
    icon = {"Implemented": "✅", "Open": "❌", "Unknown": "⚠️"}
    for r in rows:
        lines.append(f"| {r['id']} | {r['title']} | {icon[r['status']]} {r['status']} | "
                     f"{r['iso']} | {r['soc2']} | `{r['evidence']}` | {r['remediation']} |")
    lines.append("")
    if changed:
        lines.append("## ⚠ Drift detected — the compliance documents may need review\n")
        lines.append(f"Monitored source files changed since last sync: {', '.join('`'+c+'`' for c in changed)}\n")
        lines.append(f"**Affected controls:** {', '.join(affected) or 'n/a'}\n")
        lines.append("**REVIEW REQUIRED** in these documents (verify narrative + tables still match code):")
        for d in THE_SIX_DOCS:
            lines.append(f"- [ ] {d}")
        lines.append("")
    else:
        lines.append("_No monitored source changes since last sync — documents considered current._\n")
    STATUS_MD.write_text("\n".join(lines), encoding="utf-8")

def write_status_csv(rows, head, stamp):
    import csv
    with STATUS_CSV.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["commit", "generated", "id", "title", "status", "iso_27001", "soc2_tsc", "evidence", "remediation"])
        for r in rows:
            w.writerow([head, stamp, r["id"], r["title"], r["status"], r["iso"], r["soc2"], r["evidence"], r["remediation"]])

def update_xlsx(rows, head, stamp):
    if not EVIDENCE_XLSX.exists():
        return False
    try:
        import openpyxl
    except ImportError:
        return False
    wb = openpyxl.load_workbook(EVIDENCE_XLSX)
    name = "Live Control Status"
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name)
    ws.append([f"Auto-generated · commit {head} · {stamp}"])
    ws.append(["ID", "Control", "Status", "ISO 27001:2022", "SOC 2 TSC", "Evidence", "Remediation"])
    for r in rows:
        ws.append([r["id"], r["title"], r["status"], r["iso"], r["soc2"], r["evidence"], r["remediation"]])
    wb.save(EVIDENCE_XLSX)
    return True

def _merged_prs():
    """Merged PRs via `gh` (preferred), else parse merge-commits from `git log`.
    Returns list of dicts: number, title, author, createdAt, mergedAt, head, oid."""
    try:
        out = subprocess.run(
            ["gh", "pr", "list", "--state", "merged", "--limit", "300", "--json",
             "number,title,author,createdAt,mergedAt,headRefName,mergeCommit"],
            cwd=REPO, capture_output=True, text=True, timeout=90)
        if out.returncode == 0 and out.stdout.strip():
            res = []
            for p in json.loads(out.stdout):
                res.append(dict(
                    number=p.get("number"), title=(p.get("title") or "").strip(),
                    author=(p.get("author") or {}).get("login", "") or "",
                    createdAt=(p.get("createdAt") or "")[:10],
                    mergedAt=(p.get("mergedAt") or "")[:10],
                    head=p.get("headRefName", "") or "",
                    oid=((p.get("mergeCommit") or {}).get("oid", "") or "")[:8]))
            return res
    except Exception:
        pass
    # fallback: git log merge commits ("Merge pull request #N ...")
    try:
        out = subprocess.run(["git", "log", "--merges", "--date=short",
                              "--pretty=%h|%ad|%an|%s", "-n", "300"],
                             cwd=REPO, capture_output=True, text=True)
        res = []
        for line in out.stdout.splitlines():
            parts = line.split("|", 3)
            if len(parts) < 4:
                continue
            h, d, an, subj = parts
            m = re.search(r"Merge pull request #(\d+)", subj)
            if not m:
                continue
            res.append(dict(number=int(m.group(1)), title=subj, author=an,
                            createdAt="", mergedAt=d, head="", oid=h))
        return res
    except Exception:
        return []

def append_merged_prs():
    """Idempotently append merged PRs not already referenced in the SDLC Change Log.
    Returns count appended, or -1 if skipped (no xlsx / no openpyxl / no sheet)."""
    if not SDLC_XLSX.exists():
        return -1
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Border, Side, PatternFill
    except ImportError:
        return -1
    prs = _merged_prs()
    if not prs:
        return 0
    wb = openpyxl.load_workbook(SDLC_XLSX)
    if "Change Log" not in wb.sheetnames:
        return -1
    ws = wb["Change Log"]
    # PR numbers already present anywhere in the sheet (manual rows included)
    seen = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        for v in row:
            if v is None:
                continue
            for m in re.finditer(r"#(\d+)", str(v)):
                seen.add(int(m.group(1)))
    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    green = PatternFill("solid", fgColor="DCFCE7")
    added = 0
    for p in sorted(prs, key=lambda x: (x["number"] or 0)):
        num = p["number"]
        if not num or num in seen:
            continue
        impl = (f"{p['oid']} ({p['head']})" if p["oid"] and p["head"]
                else p["oid"] or p["head"] or "-")
        rowvals = [f"CL-PR{num}", p["title"], "PR (merged)", p["head"] or "-",
                   p["createdAt"] or "", "Yes", "Yes", p["author"] or "-",
                   impl, "CI / review", "PR review + merge", f"PR #{num}",
                   f"Merged {p['mergedAt']}".strip()]
        ws.append(rowvals)
        rr = ws.max_row
        for c in range(1, len(rowvals) + 1):
            ws.cell(row=rr, column=c).border = border
            ws.cell(row=rr, column=c).alignment = Alignment(vertical="top", wrap_text=True)
        ws.cell(row=rr, column=13).fill = green
        seen.add(num)
        added += 1
    if added:
        wb.save(SDLC_XLSX)
    return added

def main():
    mode = "update"
    if "--check" in sys.argv: mode = "check"
    elif "--update" not in sys.argv and len(sys.argv) > 1: mode = "dry"
    elif "--update" not in sys.argv: mode = "dry"

    DOCS.mkdir(parents=True, exist_ok=True)
    head = git_head()
    stamp = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    rows = run_checks()
    cur, changed, affected = detect_drift()

    openc = [r for r in rows if r["status"] == "Open"]
    print(f"[compliance-sync] commit {head} · {len(rows)-len(openc)}/{len(rows)} Implemented, {len(openc)} Open")
    for r in rows:
        print(f"  {r['id']:<4} {r['status']:<12} {r['title']}")
    if changed:
        print(f"[compliance-sync] DRIFT: changed {changed} → affected controls {affected}")

    if mode in ("update",):
        write_status_md(rows, changed, affected, head, stamp)
        write_status_csv(rows, head, stamp)
        xlsx = update_xlsx(rows, head, stamp)
        pr_added = append_merged_prs()
        MANIFEST.write_text(json.dumps({"commit": head, "generated": stamp, "files": cur}, indent=2))
        print(f"[compliance-sync] wrote {STATUS_MD.name}, {STATUS_CSV.name}"
              + (f", updated {EVIDENCE_XLSX.name}" if xlsx else "")
              + (f", appended {pr_added} merged PR(s) to SDLC Change Log" if pr_added and pr_added > 0
                 else (", SDLC Change Log up to date" if pr_added == 0 else "")))
        if changed:
            print("[compliance-sync] Documents flagged for REVIEW (see COMPLIANCE-STATUS.md).")

    if mode == "check":
        # CI gate: fail if drift is detected but the docs weren't refreshed in this commit
        if changed:
            print("::error::Compliance drift detected — regenerate docs "
                  "(`python3 scripts/compliance/compliance_sync.py --update`) and commit.")
            sys.exit(1)
    print("[compliance-sync] done.")

if __name__ == "__main__":
    main()
